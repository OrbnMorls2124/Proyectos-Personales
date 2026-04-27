"""
Utilidades para generación de PDFs y reportes.
"""
from io import BytesIO
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    PageBreak, Image
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from apps.billing.models import Invoice


class InvoicePDFGenerator:
    """Generador de PDFs para facturas."""
    
    def __init__(self, invoice: Invoice):
        self.invoice = invoice
        self.buffer = BytesIO()
    
    def generate(self) -> BytesIO:
        """Genera el PDF de la factura."""
        doc = SimpleDocTemplate(
            self.buffer,
            pagesize=A4,
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch
        )
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Contenido del documento
        elements = []
        
        # Encabezado
        elements.append(Paragraph("FACTURA", title_style))
        elements.append(Spacer(1, 0.2 * inch))
        
        # Información general
        info_data = [
            ['Número de Factura:', self.invoice.invoice_number],
            ['Fecha:', self.invoice.invoice_date.strftime('%d/%m/%Y')],
            ['Fecha Vencimiento:', self.invoice.due_date.strftime('%d/%m/%Y')],
            ['Estado:', self.invoice.get_status_display()],
        ]
        info_table = Table(info_data, colWidths=[2*inch, 2*inch])
        info_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3 * inch))
        
        # Datos del cliente
        elements.append(Paragraph("<b>Información del Cliente</b>", styles['Heading2']))
        client_data = [
            ['Nombre:', self.invoice.customer.name],
            ['Cédula/NIT:', self.invoice.customer.id_number],
            ['Email:', self.invoice.customer.email],
            ['Teléfono:', self.invoice.customer.phone],
            ['Dirección:', self.invoice.customer.get_full_address()],
        ]
        client_table = Table(client_data, colWidths=[1.5*inch, 4*inch])
        client_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ]))
        elements.append(client_table)
        elements.append(Spacer(1, 0.3 * inch))
        
        # Líneas de factura
        elements.append(Paragraph("<b>Productos/Servicios</b>", styles['Heading2']))
        items_data = [['Producto', 'Cantidad', 'Precio Unitario', 'Descuento %', 'Total']]
        
        for item in self.invoice.invoice_items.all():
            items_data.append([
                item.product.name,
                str(item.quantity),
                f"${item.unit_price:.2f}",
                f"{item.discount_percent}%",
                f"${item.get_total():.2f}"
            ])
        
        items_table = Table(items_data, colWidths=[2.5*inch, 0.8*inch, 1.2*inch, 0.9*inch, 1.2*inch])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 8),
        ]))
        elements.append(items_table)
        elements.append(Spacer(1, 0.3 * inch))
        
        # Totales
        totals_data = [
            ['Subtotal:', f"${self.invoice.subtotal:.2f}"],
            ['Impuesto ({0}%):'.format(self.invoice.tax_rate), f"${self.invoice.tax_amount:.2f}"],
            ['Descuento:', f"-${self.invoice.discount_amount:.2f}"],
            ['TOTAL:', f"${self.invoice.total:.2f}"],
            ['Pagado:', f"${self.invoice.paid_amount:.2f}"],
            ['Pendiente:', f"${self.invoice.get_remaining_balance():.2f}"],
        ]
        
        totals_table = Table(totals_data, colWidths=[4*inch, 2*inch])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 9),
        ]))
        elements.append(totals_table)
        
        # Método de pago y notas
        if self.invoice.payment_method:
            elements.append(Spacer(1, 0.2 * inch))
            elements.append(Paragraph(
                f"<b>Método de Pago:</b> {self.invoice.get_payment_method_display()}",
                styles['Normal']
            ))
        
        if self.invoice.notes:
            elements.append(Spacer(1, 0.2 * inch))
            elements.append(Paragraph(
                f"<b>Notas:</b> {self.invoice.notes}",
                styles['Normal']
            ))
        
        # Pie de página
        elements.append(Spacer(1, 0.5 * inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        elements.append(Paragraph(
            f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
            footer_style
        ))
        
        # Construir PDF
        doc.build(elements)
        self.buffer.seek(0)
        return self.buffer
    
    def get_filename(self) -> str:
        """Retorna el nombre del archivo."""
        return f"factura_{self.invoice.invoice_number}.pdf"


class ReportExcelGenerator:
    """Generador de reportes en Excel."""
    
    @staticmethod
    def generate_sales_report(start_date, end_date):
        """Genera reporte de ventas en Excel."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from apps.billing.models import Invoice
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas"
        
        # Encabezados
        headers = ['Factura', 'Cliente', 'Fecha', 'Total', 'Impuesto', 'Pagado', 'Estado']
        ws.append(headers)
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Datos
        invoices = Invoice.objects.filter(
            invoice_date__range=[start_date, end_date]
        )
        
        for invoice in invoices:
            ws.append([
                invoice.invoice_number,
                invoice.customer.name,
                invoice.invoice_date.strftime('%d/%m/%Y'),
                f"${invoice.total:.2f}",
                f"${invoice.tax_amount:.2f}",
                f"${invoice.paid_amount:.2f}",
                invoice.get_status_display(),
            ])
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        
        # Guardar
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
